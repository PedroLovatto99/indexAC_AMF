import os
from openpyxl import load_workbook
from django.conf import settings

def gerar_excel_final(dados_ia, nome_aluno, curso_aluno):
    caminho_modelo = os.path.join(settings.BASE_DIR, 'doc', 'Horas complementares.xlsx')
    
    wb = load_workbook(caminho_modelo)
    planilha = wb.active

    
    planilha['A7'] = f"Nome completo aluno: {nome_aluno}"
    planilha['A8'] = f"Curso: {curso_aluno}"

    linha_inicio = 12
    
    for indice, certificado in enumerate(dados_ia):
        linha_atual = linha_inicio + indice
        

        planilha.cell(row=linha_atual, column=1, value=certificado.get('evento', ''))
        planilha.cell(row=linha_atual, column=2, value=certificado.get('horas', ''))
        planilha.cell(row=linha_atual, column=3, value=certificado.get('data', ''))

    nome_arquivo_novo = f"Horas Complementares.xlsx"
    caminho_salvo = os.path.join(settings.BASE_DIR, 'media', nome_arquivo_novo)
    
    wb.save(caminho_salvo)
    
    return caminho_salvo